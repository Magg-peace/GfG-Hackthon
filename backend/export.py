"""Export functionality: generate PDF reports and Excel workbooks from dashboard data.

PDF  → fpdf2  (embedded matplotlib chart images + styled tables)
Excel → openpyxl (data sheets + native bar/line/pie charts per result)
"""

from __future__ import annotations

import io
import json
import math
import re
from datetime import datetime
from typing import Any

# ── Optional heavy imports (fail gracefully) ──────────────────────────────────
try:
    import matplotlib
    matplotlib.use("Agg")          # headless backend
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    from matplotlib.figure import Figure
    _MATPLOTLIB = True
except ImportError:
    _MATPLOTLIB = False

try:
    from fpdf import FPDF
    _FPDF = True
except ImportError:
    _FPDF = False

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                 Border, Side, GradientFill)
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ── Colour palette ─────────────────────────────────────────────────────────────

_PALETTE = [
    "#3b82f6", "#22c55e", "#f59e0b", "#ef4444", "#8b5cf6",
    "#06b6d4", "#ec4899", "#f97316", "#14b8a6", "#a855f7",
]


def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))  # type: ignore


# ── Chart rendering helpers (matplotlib) ──────────────────────────────────────

def _render_chart_to_png(chart: dict) -> bytes | None:
    """Render a chart config + data dict to a PNG image (bytes). Returns None on failure."""
    if not _MATPLOTLIB:
        return None

    data = chart.get("data", [])
    if not data:
        return None

    chart_type = chart.get("chart_type", "bar")
    title = chart.get("title", "")
    x_col = chart.get("x_axis", "")
    y_cols = chart.get("y_axis", [])
    x_label = chart.get("x_label", x_col)
    y_label = chart.get("y_label", ", ".join(y_cols))

    try:
        fig, ax = plt.subplots(figsize=(8, 4.5), dpi=110)
        fig.patch.set_facecolor("#1e293b")
        ax.set_facecolor("#1e293b")
        ax.tick_params(colors="white", labelsize=8)
        ax.xaxis.label.set_color("white")
        ax.yaxis.label.set_color("white")
        ax.title.set_color("white")
        for spine in ax.spines.values():
            spine.set_edgecolor("#334155")

        if chart_type == "metric":
            # Render as a big number
            val_col = chart.get("value_column", y_cols[0] if y_cols else "")
            val = data[0].get(val_col, "N/A") if data else "N/A"
            prefix = chart.get("prefix", "")
            suffix = chart.get("suffix", "")
            ax.text(0.5, 0.5, f"{prefix}{val}{suffix}",
                    ha="center", va="center", fontsize=36,
                    color="#3b82f6", fontweight="bold",
                    transform=ax.transAxes)
            ax.axis("off")

        elif chart_type == "pie":
            if not x_col or not y_cols:
                plt.close(fig)
                return None
            labels = [str(d.get(x_col, "")) for d in data[:10]]
            vals = []
            for d in data[:10]:
                try:
                    vals.append(float(d.get(y_cols[0], 0) or 0))
                except (ValueError, TypeError):
                    vals.append(0)
            colors = _PALETTE[:len(labels)]
            ax.pie(vals, labels=labels, autopct="%1.1f%%", colors=colors,
                   textprops={"color": "white", "fontsize": 8})

        elif chart_type in ("bar",):
            if not x_col or not y_cols:
                plt.close(fig)
                return None
            x_vals = [str(d.get(x_col, "")) for d in data]
            width = 0.8 / max(len(y_cols), 1)
            for i, ycol in enumerate(y_cols):
                y_vals = []
                for d in data:
                    try:
                        y_vals.append(float(d.get(ycol, 0) or 0))
                    except (ValueError, TypeError):
                        y_vals.append(0)
                offset = (i - len(y_cols) / 2 + 0.5) * width
                xs = [j + offset for j in range(len(x_vals))]
                ax.bar(xs, y_vals, width=width,
                       color=_PALETTE[i % len(_PALETTE)], label=ycol, alpha=0.85)
            ax.set_xticks(range(len(x_vals)))
            ax.set_xticklabels(x_vals, rotation=35, ha="right", fontsize=7)
            ax.set_xlabel(x_label, color="white", fontsize=9)
            ax.set_ylabel(y_label, color="white", fontsize=9)
            if len(y_cols) > 1:
                ax.legend(fontsize=8, labelcolor="white",
                          facecolor="#1e293b", edgecolor="#334155")

        elif chart_type in ("line", "area"):
            if not x_col or not y_cols:
                plt.close(fig)
                return None
            x_vals = [str(d.get(x_col, "")) for d in data]
            xs = range(len(x_vals))
            for i, ycol in enumerate(y_cols):
                y_vals = []
                for d in data:
                    try:
                        y_vals.append(float(d.get(ycol, 0) or 0))
                    except (ValueError, TypeError):
                        y_vals.append(0)
                color = _PALETTE[i % len(_PALETTE)]
                if chart_type == "area":
                    ax.fill_between(list(xs), y_vals, alpha=0.25, color=color)
                ax.plot(list(xs), y_vals, marker="o", markersize=3,
                        color=color, linewidth=1.8, label=ycol)
            step = max(1, len(x_vals) // 10)
            ax.set_xticks(list(xs)[::step])
            ax.set_xticklabels(x_vals[::step], rotation=35, ha="right", fontsize=7)
            ax.set_xlabel(x_label, color="white", fontsize=9)
            ax.set_ylabel(y_label, color="white", fontsize=9)
            if len(y_cols) > 1:
                ax.legend(fontsize=8, labelcolor="white",
                          facecolor="#1e293b", edgecolor="#334155")

        elif chart_type == "scatter":
            if len(y_cols) < 1:
                plt.close(fig)
                return None
            y2 = y_cols[1] if len(y_cols) > 1 else y_cols[0]
            xs2, ys2 = [], []
            for d in data:
                try:
                    xs2.append(float(d.get(x_col, 0) or 0))
                    ys2.append(float(d.get(y_cols[0], 0) or 0))
                except (ValueError, TypeError):
                    pass
            ax.scatter(xs2, ys2, color=_PALETTE[0], alpha=0.7, s=20)
            ax.set_xlabel(x_label, color="white", fontsize=9)
            ax.set_ylabel(y_label, color="white", fontsize=9)

        else:
            plt.close(fig)
            return None

        ax.set_title(title, color="white", fontsize=10, pad=8)
        ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda x, _: f"{x/1e6:.1f}M" if abs(x) >= 1e6
                                  else (f"{x/1e3:.1f}K" if abs(x) >= 1e3 else f"{x:g}"))
        )
        fig.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        plt.close(fig)
        buf.seek(0)
        return buf.read()

    except Exception:
        try:
            plt.close(fig)
        except Exception:
            pass
        return None


# ── PDF export ─────────────────────────────────────────────────────────────────

class _BIDashboardPDF(FPDF):  # type: ignore
    def __init__(self, title: str = "BI Dashboard Report"):
        super().__init__()
        self._report_title = title
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_fill_color(30, 41, 59)   # slate-800
        self.rect(0, 0, 210, 18, "F")
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(255, 255, 255)
        self.set_xy(10, 4)
        self.cell(0, 10, self._report_title, align="L")
        self.set_font("Helvetica", "", 8)
        self.set_xy(-60, 4)
        self.cell(50, 10, datetime.now().strftime("%Y-%m-%d %H:%M"), align="R")
        self.ln(12)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 8, f"Page {self.page_no()}", align="C")


def export_to_pdf(
    charts: list[dict],
    summary: str,
    query: str = "",
    session_id: str = "",
) -> bytes:
    """Generate a PDF report from chart configs + data.  Returns raw bytes."""
    if not _FPDF:
        raise ImportError("fpdf2 is not installed. Run: pip install fpdf2")

    pdf = _BIDashboardPDF(title="BI Dashboard – Analysis Report")
    pdf.add_page()

    # ── Cover section ──────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(59, 130, 246)   # blue
    pdf.cell(0, 10, "Analysis Report", ln=True)
    pdf.ln(2)

    if query:
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(100, 116, 139)
        pdf.multi_cell(0, 5, f"Query: {query}")
        pdf.ln(2)

    if summary:
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(30, 41, 59)
        pdf.set_fill_color(248, 250, 252)
        pdf.set_draw_color(226, 232, 240)
        pdf.multi_cell(0, 6, summary, border=1, fill=True)
        pdf.ln(4)

    # ── Charts ─────────────────────────────────────────────────────────────────
    for idx, chart in enumerate(charts):
        if chart.get("error"):
            continue

        chart_type = chart.get("chart_type", "bar")
        title = chart.get("title", f"Chart {idx+1}")
        insight = chart.get("insight", "")
        data = chart.get("data", [])

        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(0, 8, f"{idx+1}. {title}", ln=True)

        if insight:
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(100, 116, 139)
            pdf.multi_cell(0, 5, f"Insight: {insight}")

        # Try to render chart image
        if chart_type != "table" and _MATPLOTLIB:
            img_bytes = _render_chart_to_png(chart)
            if img_bytes:
                buf = io.BytesIO(img_bytes)
                # Check if we have space
                if pdf.get_y() + 75 > pdf.h - 20:
                    pdf.add_page()
                pdf.image(buf, x=10, w=190, h=0)
                pdf.ln(4)

        # Data table (always for "table" type; optional summary for others)
        if data and (chart_type == "table" or chart_type == "metric" or not _MATPLOTLIB):
            _draw_pdf_table(pdf, data, max_rows=30)
            pdf.ln(4)

        if pdf.get_y() > pdf.h - 40:
            pdf.add_page()

    # ── Per-chart data pages ───────────────────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 8, "Filtered / Detailed Data", ln=True)
    pdf.ln(2)

    for idx, chart in enumerate(charts):
        data = chart.get("data", [])
        if not data:
            continue
        title = chart.get("title", f"Dataset {idx+1}")
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(59, 130, 246)
        pdf.cell(0, 7, title, ln=True)
        _draw_pdf_table(pdf, data, max_rows=100)
        pdf.ln(5)
        if pdf.get_y() > pdf.h - 40:
            pdf.add_page()

    return bytes(pdf.output())


def _draw_pdf_table(pdf: "FPDF", data: list[dict], max_rows: int = 50) -> None:
    if not data:
        return
    cols = list(data[0].keys())
    col_w = min(40, int(190 / max(len(cols), 1)))

    # Header row
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    for col in cols:
        label = str(col)[:18]
        pdf.cell(col_w, 5, label, border=1, fill=True, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 7)
    for i, row in enumerate(data[:max_rows]):
        if i % 2 == 0:
            pdf.set_fill_color(248, 250, 252)
        else:
            pdf.set_fill_color(241, 245, 249)
        pdf.set_text_color(30, 41, 59)
        for col in cols:
            val = str(row.get(col, ""))[:20]
            pdf.cell(col_w, 4, val, border=1, align="L")
        pdf.ln()

    if len(data) > max_rows:
        pdf.set_font("Helvetica", "I", 7)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 5, f"  … {len(data) - max_rows} more rows not shown.", ln=True)


# ── Excel export ───────────────────────────────────────────────────────────────

_HEADER_FILL = "1E293B"
_ALT_FILL_1  = "F8FAFC"
_ALT_FILL_2  = "F1F5F9"
_BLUE        = "3B82F6"


def export_to_excel(
    charts: list[dict],
    summary: str,
    query: str = "",
    session_id: str = "",
) -> bytes:
    """Generate an Excel workbook. Returns raw bytes."""
    if not _OPENPYXL:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Analysis Summary"
    _apply_summary_sheet(ws_summary, query, summary, charts)

    # ── One data sheet + chart per result ──────────────────────────────────────
    for idx, chart_cfg in enumerate(charts, 1):
        data = chart_cfg.get("data", [])
        if not data:
            continue
        sheet_name = _safe_sheet_name(chart_cfg.get("title", f"Chart {idx}"), idx)
        ws = wb.create_sheet(title=sheet_name)
        _fill_data_sheet(ws, chart_cfg, data)
        _add_excel_chart(ws, chart_cfg, data, idx)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _safe_sheet_name(name: str, idx: int) -> str:
    safe = re.sub(r"[\\/*?:\[\]]", "", name)[:28]
    return safe or f"Sheet{idx}"


def _apply_summary_sheet(ws, query: str, summary: str, charts: list[dict]) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    label_font  = Font(name="Calibri", bold=True, size=11, color="1E293B")
    body_font   = Font(name="Calibri", size=10, color="374151")

    def _hdr(row, text):
        ws.cell(row, 1, text).font = header_font
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=_HEADER_FILL)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 22

    ws.cell(1, 1, "BI Dashboard – Analysis Report").font = Font(
        name="Calibri", bold=True, size=18, color=_BLUE)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 32

    ws.cell(3, 1, "Generated:").font = label_font
    ws.cell(3, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = body_font

    ws.cell(4, 1, "Query:").font = label_font
    ws.cell(4, 2, query or "—").font = body_font
    ws.cell(4, 2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 30

    ws.cell(6, 1, "Summary:").font = label_font
    ws.cell(6, 2, summary or "—").font = body_font
    ws.cell(6, 2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = max(15 * math.ceil(len(summary or "") / 80), 30)

    ws.cell(8, 1, "Charts in this report:").font = label_font
    for i, c in enumerate(charts, 1):
        ws.cell(8 + i, 1, f"  {i}.").font = body_font
        ws.cell(8 + i, 2, c.get("title", "")).font = body_font


def _fill_data_sheet(ws, chart_cfg: dict, data: list[dict]) -> None:
    if not data:
        return
    cols = list(data[0].keys())

    header_font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor=_HEADER_FILL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Insight + title row
    ws.cell(1, 1, chart_cfg.get("title", "")).font = Font(
        name="Calibri", bold=True, size=13, color=_BLUE)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(len(cols), 1))

    insight = chart_cfg.get("insight", "")
    if insight:
        ws.cell(2, 1, insight).font = Font(name="Calibri", size=9, color="64748B", italic=True)
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=max(len(cols), 1))

    start_row = 4

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(start_row, ci, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) + 4)

    for di, row in enumerate(data, start_row + 1):
        fill = PatternFill("solid", fgColor=_ALT_FILL_1 if di % 2 == 0 else _ALT_FILL_2)
        for ci, col in enumerate(cols, 1):
            val = row.get(col)
            if val is not None:
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    val = str(val)
            cell = ws.cell(di, ci, val)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=9)

    ws.freeze_panes = ws.cell(start_row + 1, 1)


def _add_excel_chart(ws, chart_cfg: dict, data: list[dict], sheet_idx: int) -> None:
    """Append a native Excel chart to *ws* using openpyxl chart objects."""
    if not data:
        return

    chart_type = chart_cfg.get("chart_type", "bar")
    x_col = chart_cfg.get("x_axis", "")
    y_cols = chart_cfg.get("y_axis", [])
    cols = list(data[0].keys())

    start_row = 4   # header is at row 4 (see _fill_data_sheet)
    data_start = start_row + 1
    data_end   = data_start + len(data) - 1

    chart_anchor_row = data_end + 3

    def _col_index(col_name: str) -> int | None:
        try:
            return cols.index(col_name) + 1
        except ValueError:
            return None

    if chart_type in ("bar",):
        ec = BarChart()
        ec.type = "col"
        ec.title = chart_cfg.get("title", "")
        ec.y_axis.title = chart_cfg.get("y_label", "")
        ec.x_axis.title = chart_cfg.get("x_label", "")
        ec.style = 10
        ec.width = 20
        ec.height = 12

        x_idx = _col_index(x_col)
        if x_idx:
            cats = Reference(ws, min_col=x_idx, min_row=data_start, max_row=data_end)
            ec.set_categories(cats)
        for ycol in (y_cols or []):
            y_idx = _col_index(ycol)
            if y_idx:
                data_ref = Reference(ws, min_col=y_idx, min_row=start_row, max_row=data_end)
                ec.add_data(data_ref, titles_from_data=True)

        anchor = f"A{chart_anchor_row}"
        ws.add_chart(ec, anchor)

    elif chart_type in ("line", "area"):
        ec = LineChart()
        ec.title = chart_cfg.get("title", "")
        ec.width = 20
        ec.height = 12
        ec.style = 12

        x_idx = _col_index(x_col)
        if x_idx:
            cats = Reference(ws, min_col=x_idx, min_row=data_start, max_row=data_end)
            ec.set_categories(cats)
        for ycol in (y_cols or []):
            y_idx = _col_index(ycol)
            if y_idx:
                data_ref = Reference(ws, min_col=y_idx, min_row=start_row, max_row=data_end)
                ec.add_data(data_ref, titles_from_data=True)

        ws.add_chart(ec, f"A{chart_anchor_row}")

    elif chart_type == "pie":
        if not y_cols:
            return
        ec = PieChart()
        ec.title = chart_cfg.get("title", "")
        ec.width = 16
        ec.height = 12

        y_idx = _col_index(y_cols[0]) if y_cols else None
        x_idx = _col_index(x_col)
        if y_idx:
            data_ref = Reference(ws, min_col=y_idx, min_row=start_row, max_row=data_end)
            ec.add_data(data_ref, titles_from_data=True)
        if x_idx:
            labels = Reference(ws, min_col=x_idx, min_row=data_start, max_row=data_end)
            ec.set_categories(labels)

        ws.add_chart(ec, f"A{chart_anchor_row}")
